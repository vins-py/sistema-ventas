import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font

ventas = pd.read_csv("ventas.csv")

total_por_dia = ventas.groupby("fecha", as_index=False)["total"].sum()

productos_por_dia = ventas.groupby(["fecha", "producto"], as_index=False).agg({
    "cantidad": "sum",
    "total": "sum"
})

productos_totales = ventas.groupby("producto")["cantidad"].sum()

total_general = ventas["total"].sum()
cantidad_total = ventas["cantidad"].sum()
producto_mas_vendido = productos_totales.idxmax()
cantidad_producto_mas_vendido = productos_totales.max()

resumen = pd.DataFrame({
    "Dato": [
        "Total vendido",
        "Cantidad total vendida",
        "Producto más vendido",
        "Cantidad del producto más vendido"
    ],
    "Valor": [
        total_general,
        cantidad_total,
        producto_mas_vendido,
        cantidad_producto_mas_vendido
    ]
})

nombre_reporte = "reporte_ventas.xlsx"

with pd.ExcelWriter(nombre_reporte, engine="openpyxl") as writer:
    resumen.to_excel(writer, sheet_name="Resumen", index=False)
    ventas.to_excel(writer, sheet_name="Ventas", index=False)
    total_por_dia.to_excel(writer, sheet_name="Total por dia", index=False)
    productos_por_dia.to_excel(writer, sheet_name="Productos por dia", index=False)

libro = load_workbook(nombre_reporte)

for hoja in libro.worksheets:
    hoja.freeze_panes = "A2"

    for celda in hoja[1]:
        celda.font = Font(bold=True)

    for columna in hoja.columns:
        ancho_maximo = 0
        letra_columna = columna[0].column_letter

        for celda in columna:
            if celda.value is not None:
                ancho_maximo = max(ancho_maximo, len(str(celda.value)))

        hoja.column_dimensions[letra_columna].width = ancho_maximo + 3

    for fila in hoja.iter_rows(min_row=2):
        for celda in fila:
            encabezado = hoja.cell(row=1, column=celda.column).value

            if encabezado in ["precio", "subtotal", "iva", "total"]:
                celda.number_format = '$#,##0.00'

for celda in libro["Resumen"]["B"]:
    if isinstance(celda.value, float):
        celda.number_format = '$#,##0.00'

libro.save(nombre_reporte)

print("Reporte con resumen creado: reporte_ventas.xlsx")